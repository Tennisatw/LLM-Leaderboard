import pandas as pd, numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

inp = "leaderboard_full.xlsx"
out = "leaderboard_pretty.xlsx"

df = pd.read_excel(inp)

# 你想在Excel里显示的列名 -> 原始文件里的列名（或你内部标准名）
rename_map = {
    "总排名": "ranking",
    "模型名称": "canonical",
    "LMArene": "lmarena_text",
    "MMMU Pro": "mmmupro",
    "GPQA Diamond": "gpqadiamond",
    "HLE": "hle",
    "OmniScience": "omniscience",
    "Long Context Reasoning": "aalcr",
    "Instruction Follow Bench": "ifbench",
    "TerminalBench": "terminalbenchhard",
    "Scicode": "scicode",
    "GDPval": "gdpvalaa",
}

# 先把原始列名映射成你想显示的列名（方便后面全部按你想要的名字写）
inv = {v: k for k, v in rename_map.items() if v in df.columns}
df = df.rename(columns=inv)
df = df.rename(columns={"weighted_score": "总分"})

left_cols = ["总排名", "模型名称", "总分"]
groups = [
    ("人类偏好", ["LMArene"], "1f4e79"),
    ("知识与推理", ["MMMU Pro", "GPQA Diamond", "HLE", "OmniScience"], "9c4a00"),
    ("长文本推理", ["Long Context Reasoning"], "a67c00"),
    ("指令遵循", ["Instruction Follow Bench"], "2f6b2f"),
    ("代码能力", ["TerminalBench", "Scicode"], "6a1b9a"),
    ("真实世界任务", ["GDPval"], "424242"),
]

cols = [c for c in left_cols if c in df.columns]
heat_cols = []
for _, cs, _ in groups:
    heat_cols += [c for c in cs if c in df.columns]

df = df[cols + heat_cols]
df.insert(cols.index("模型名称") + 1, "总分", df.pop("总分"))

left_cols = ["总排名", "模型名称", "总分"]
cols = [c for c in left_cols if c in df.columns]

df2 = pd.concat(
    [pd.DataFrame([[""] * len(df.columns)], columns=df.columns),
     pd.DataFrame([df.columns.tolist()], columns=df.columns),
     df],
    ignore_index=True
)

with pd.ExcelWriter(out, engine="openpyxl") as w:
    df2.to_excel(w, index=False, header=False, sheet_name="Leaderboard")

wb = load_workbook(out)
ws = wb["Leaderboard"]

no_border = Border()
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

ws.row_dimensions[1].height = 32
ws.row_dimensions[2].height = 24
for r in range(3, ws.max_row + 1):
    ws.row_dimensions[r].height = 18

font_hdr1 = Font(color="ffffff", bold=True, size=14)
font_hdr2 = Font(color="ffffff", bold=True, size=11)
font_cell = Font(color="000000", size=10)

ws.freeze_panes = "C3"

ws.column_dimensions["A"].width = 8
ws.column_dimensions["B"].width = 22
for c in range(3, ws.max_column + 1):
    ws.column_dimensions[get_column_letter(c)].width = 12

for r in range(1, ws.max_row + 1):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(r, c)
        cell.border = no_border
        cell.alignment = center
        cell.font = font_cell

fill_hdr = PatternFill("solid", fgColor="2b2b2b")
for c in range(1, ws.max_column + 1):
    cell = ws.cell(2, c)
    cell.fill = fill_hdr
    cell.font = font_hdr2
for c in range(1, len(cols) + 1):
    cell = ws.cell(1, c)
    cell.fill = fill_hdr
    cell.font = font_hdr1

col_index = {ws.cell(2, c).value: c for c in range(1, ws.max_column + 1)}
for title, cs, color in groups:
    cs = [x for x in cs if x in col_index]
    if not cs:
        continue
    a = col_index[cs[0]]
    b = col_index[cs[-1]]
    ws.merge_cells(start_row=1, start_column=a, end_row=1, end_column=b)
    cell = ws.cell(1, a)
    cell.value = title
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = font_hdr1
    cell.alignment = center

if "模型名称" in col_index:
    c_can = col_index["模型名称"]
    for r in range(3, ws.max_row + 1):
        ws.cell(r, c_can).alignment = left

def set_fmt(col, fmt):
    if col not in col_index:
        return
    c = col_index[col]
    for r in range(3, ws.max_row + 1):
        ws.cell(r, c).number_format = fmt

set_fmt("总排名", "0")
set_fmt("LMArene", "0")
set_fmt("GDPval", "0")
set_fmt("总分", "0.0")

for c in heat_cols:
    if c not in ["LMArene", "GDPval"]:
        set_fmt(c, "0.0")

fill_na = PatternFill("solid", fgColor="f0f0f0")
for c in heat_cols:
    if c not in col_index:
        continue
    cc = col_index[c]
    for r in range(3, ws.max_row + 1):
        v = ws.cell(r, cc).value
        if v is None or (isinstance(v, float) and np.isnan(v)):
            ws.cell(r, cc).fill = fill_na

rule = ColorScaleRule(
    start_type="min", start_color="f8696b",
    mid_type="percentile", mid_value=50, mid_color="ffe08a",
    end_type="max", end_color="63be7b",
)
for c in heat_cols:
    cc = col_index[c]
    rng = f"{ws.cell(3, cc).coordinate}:{ws.cell(ws.max_row, cc).coordinate}"
    ws.conditional_formatting.add(rng, rule)

wb.save(out)
